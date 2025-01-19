import logging
import mimetypes
import os
import tempfile
from typing import Dict
from pypdf import PdfReader
import pytesseract
from PIL import Image
from docx import Document
from odf.opendocument import load
from odf import teletype
from ebooklib import epub
from openpyxl import load_workbook
from config.utils import get_whisper_model_and_pipeline
import xml.etree.ElementTree as ET
import json
import yaml
import wave
import tarfile
import zipfile
import py7zr
import csv
import html
import shutil
from datasets import load_dataset
import torch
from core.memory_engine import MemoryEngine

# Logger setup
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class FileParser:
    def __init__(self, memory_engine: MemoryEngine):
        """
        Initialize the FileParser class with safe extensions, AI models, and MemoryEngine.

        :param memory_engine: An instance of MemoryEngine for storing parsed content.
        """
        self.memory_engine = memory_engine
        self.safe_extensions = {
            'txt', 'pdf', 'jpg', 'png', 'jpeg', 'docx', 'csv', 'xml', 'html', 'odt', 
            'json', 'yaml', 'xlsx', 'xls', 'ods', 'mp3', 'wav', 'ogg', 'flac', 'm4a', 'epub'
        }

        # Whisper model initialization using HuggingFace transformers
        device = "cuda:0" if torch.cuda.is_available() else "cpu"
        torch_dtype = torch.float16 if torch.cuda.is_available() else torch.float32
        try:
            self.model, self.processor, self.pipeline = get_whisper_model_and_pipeline()
            logger.info("[INIT] Whisper model initialized successfully.")
        except Exception as e:
            logger.critical(f"[INIT FAILED] Failed to initialize Whisper model: {e}")
            raise RuntimeError("Failed to initialize Whisper model.")

    def parse(self, filepath, mime_type=None, language="eng"):
        logger.info(f"Attempting to parse file: {filepath}")
        """
        Parse the file based on its MIME type.

        :param filepath: Path to the file.
        :param mime_type: MIME type of the file.
        :param language: Language for OCR or transcription.
        :return: Parsed content structure.
        """
        try:
            if mime_type is None:
                mime_type, _ = mimetypes.guess_type(filepath)
            if mime_type is None:
                raise ValueError(f"Could not determine MIME type for file: {filepath}")
            if mime_type == 'txt':
                with open(filepath, 'r', encoding='utf-8') as file:
                    content = file.read()
                    # Simple chunking by paragraphs or by a fixed number of characters
                    chunks = self._chunk_content(content)
                    return {'chunks': chunks}
            if mime_type.startswith("text"):
                return self.parse_text(filepath)
            elif mime_type == "application/pdf":
                return self.parse_pdf(filepath, language)
            elif mime_type.startswith("image"):
                return self.parse_image(filepath, language)
            elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                return self.parse_docx(filepath)
            elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                return self.parse_excel(filepath)
            elif mime_type == "application/x-yaml":
                return self.parse_yaml(filepath)
            elif mime_type == "application/zip":
                if filepath.lower().endswith('.epub'):
                    return self.parse_epub(filepath)
                return self.parse_archive(filepath, 'zip', language)
            elif mime_type == "application/x-7z-compressed":
                return self.parse_archive(filepath, '7z', language)
            else:
                return self._parse_file(filepath, mime_type, language)
        except Exception as e:
            logger.error(f"Failed to parse file {filepath}. Error details: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Error parsing file {filepath}: {str(e)}")
            raise

    def _parse_file(self, filepath, mime_type=None, language="eng"):
        """
        Parse individual files based on MIME type.

        :param filepath: Path to the file.
        :param mime_type: MIME type of the file.
        :param language: Language for OCR or transcription.
        :return: Parsed content structure.
        """
        if mime_type is None:
            mime_type, _ = mimetypes.guess_type(filepath)

        if mime_type is None:
            raise ValueError(f"Could not determine MIME type for file: {filepath}")

        if mime_type.startswith("text"):
            return self.parse_text(filepath)
        elif mime_type == "application/pdf":
            return self.parse_pdf(filepath, language)
        elif mime_type.startswith("image"):
            return self.parse_image(filepath, language)
        elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return self.parse_docx(filepath)
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return self.parse_excel(filepath)
        elif mime_type == "application/x-yaml":
            return self.parse_yaml(filepath)
        elif mime_type.startswith("audio"):
            return self.parse_audio(filepath, language)
        else:
            raise ValueError(f"Unsupported file type: {mime_type}")
        
    def _chunk_content(self, content: str) -> Dict:
        """
        Chunk the content into manageable pieces.

        :param content: The content to chunk.
        :return: Dictionary with chunk types and content.
        """
        # Example: Chunk by paragraphs
        paragraphs = content.split('\n\n')
        chunks = {}
        for i, paragraph in enumerate(paragraphs, start=1):
            if paragraph.strip():  # Ensure we're not adding empty chunks
                chunks[f'paragraph_{i}'] = {
                    'name': f'Paragraph {i}',
                    'content': paragraph.strip(),
                    'sentences': paragraph.strip().split('. ')
                }
        
        return chunks if chunks else {'error': 'No valid content found in the file.'}

    def parse_audio(self, filepath, language="en"):
        """
        Parse audio files using the Whisper pipeline.

        :param filepath: Path to the audio file.
        :param language: Language for transcription.
        :return: Transcribed text.
        """
        try:
            result = self.pipeline(filepath)
            return result["text"]
        except Exception as e:
            logger.error(f"Audio parsing failed for {filepath}: {e}")
            return "Audio parsing failed."

    def parse_archive(self, filepath, archive_type, language="eng"):
        """
        Extract and process safe items from archive files.

        :param filepath: Path to the archive file.
        :param archive_type: Type of archive ('tar', 'zip', or '7z').
        :param language: Language for OCR and speech recognition.
        :return: List of parsed contents from safe files within the archive.
        """
        results = []
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                if archive_type == 'tar':
                    archive = tarfile.open(filepath, 'r:*')
                    members = archive.getmembers()
                elif archive_type == 'zip':
                    archive = zipfile.ZipFile(filepath, 'r')
                    members = archive.namelist()
                elif archive_type == '7z':
                    archive = py7zr.SevenZipFile(filepath, 'r')
                    members = archive.getnames()
                
                for member in members:
                    if not member.endswith('/'):
                        _, ext = os.path.splitext(member)
                        if ext.lower()[1:] in self.safe_extensions:
                            if archive_type == 'tar':
                                archive.extract(member, path=temp_dir)
                            else:  # zip or 7z
                                archive.extractall(path=temp_dir)

                            safe_file_path = os.path.join(temp_dir, member)
                            mime_type, _ = mimetypes.guess_type(safe_file_path)
                            content = self._parse_file(safe_file_path, mime_type, language)
                            results.append((member, content))
                        else:
                            logger.warning(f"Skipping unsafe file type: {member}")
            except Exception as e:
                logger.error(f"{archive_type.upper()} extraction failed for {filepath}: {e}")
                raise
            finally:
                if archive_type in ['tar', 'zip']:
                    archive.close()
                elif archive_type == '7z':
                    archive.close()

        return results

    def parse_epub(self, filepath, language="eng"):
        """
        Parse EPUB files.

        :param filepath: Path to the EPUB file.
        :param language: Language for OCR if needed within HTML content (default: "eng").
        :return: Extracted text from the EPUB.
        """
        try:
            book = epub.read_epub(filepath)
            content = {
                "title": book.get_metadata('DC', 'title')[0][0],
                "author": book.get_metadata('DC', 'creator')[0][0] if book.get_metadata('DC', 'creator') else "Unknown",
                "chunks": []
            }
            for item in book.get_items():
                if item.get_type() == epub.ITEM_DOCUMENT:
                    item_content = item.get_content().decode('utf-8')
                    chapter_content = self.parse_html_content(item_content)
                    content["chunks"].append({
                        "type": "chapter",
                        "content": chapter_content,
                        "sentences": chapter_content.split('. ')
                    })
            return content
        except Exception as e:
            logger.error(f"EPUB parsing failed for {filepath}: {e}")
            raise ValueError(f"EPUB parsing failed: {e}")

    def parse_html_content(self, html_content):
        """
        Parse HTML content string, which is useful for EPUB files.

        :param html_content: HTML content as a string.
        :return: Extracted text from the HTML.
        """
        try:
            return html.unescape(html_content)
        except Exception as e:
            raise ValueError(f"HTML content parsing failed: {e}")

    def parse_text(self, filepath, language=None):
        """
        Parse plain text files.

        :param filepath: Path to the text file.
        :param language: Unused here, kept for consistency.
        :return: Content structure of the text file.
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as file:
                content = file.read()
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",                 "chunks": [
                    {
                        "type": "text",
                        "content": content,
                        "sentences": content.split('. ')
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"Text parsing failed for {filepath}: {e}")

    def parse_pdf(self, filepath, language="eng"):
        """
        Parse PDF files.

        :param filepath: Path to the PDF file.
        :param language: Language for OCR if needed.
        :return: Content structure from the PDF.
        """
        try:
            with open(filepath, 'rb') as file:
                reader = PdfReader(file)
                content = ""
                for page in reader.pages:
                    content += page.extract_text()
                if not content.strip():
                    from pdf2image import convert_from_path
                    images = convert_from_path(filepath)
                    for image in images:
                        content += pytesseract.image_to_string(image, lang=language)
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "document",
                        "content": content,
                        "sentences": content.split('. ')
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"PDF parsing failed for {filepath}: {e}")

    def parse_image(self, filepath, language="eng"):
        """
        Parse image files for OCR.

        :param filepath: Path to the image file.
        :param language: Language for OCR.
        :return: Content structure from the image.
        """
        try:
            image = Image.open(filepath)
            content = pytesseract.image_to_string(image, lang=language)
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "image_text",
                        "content": content,
                        "sentences": content.split('. ')
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"Image parsing failed for {filepath}: {e}")

    def parse_docx(self, filepath, language=None):
        """
        Parse DOCX files.

        :param filepath: Path to the DOCX file.
        :param language: Unused here, kept for consistency.
        :return: Content structure from the DOCX.
        """
        try:
            doc = Document(filepath)
            content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "document",
                        "content": content,
                        "sentences": content.split('. ')
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"DOCX parsing failed for {filepath}: {e}")

    def parse_excel(self, filepath, language=None):
        """
        Parse Excel files (.xls, .xlsx).

        :param filepath: Path to the Excel file.
        :param language: Unused here, kept for consistency.
        :return: Content structure from all sheets.
        """
        try:
            workbook = load_workbook(filename=filepath, read_only=True)
            content = []
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows(values_only=True):
                    content.append(', '.join(str(cell) for cell in row if cell is not None))
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "spreadsheet",
                        "content": '\n'.join(content),
                        "sentences": ['\n'.join(content)]  # Treating spreadsheet data as one sentence
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"Excel parsing failed for {filepath}: {e}")

    def parse_yaml(self, filepath, language=None):
        """
        Parse YAML files.

        :param filepath: Path to the YAML file.
        :param language: Unused here, kept for consistency.
        :return: String representation of YAML data.
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as file:
                data = yaml.safe_load(file)
            yaml_str = yaml.dump(data, default_flow_style=False)
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "yaml",
                        "content": yaml_str,
                        "sentences": [yaml_str]  # Treating YAML data as one sentence
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"YAML parsing failed for {filepath}: {e}")

    def parse_json(self, filepath, language=None):
        """
        Parse JSON files.

        :param filepath: Path to the JSON file.
        :param language: Unused here, kept for consistency.
        :return: Parsed content structure.
        """
        try:
            with open(filepath, 'r') as file:
                content = json.load(file)
            json_str = json.dumps(content, indent=4)
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "json",
                        "content": json_str,
                        "sentences": [json_str]  # Treating JSON data as one sentence
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"JSON parsing failed for {filepath}: {e}")

    def parse_csv(self, filepath, language=None):
        """
        Parse CSV files.
    
        :param filepath: Path to the CSV file.
        :param language: Unused here, kept for consistency.
        :return: Extracted content structure.
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                content = '\n'.join([', '.join(row) for row in reader])
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "csv",
                        "content": content,
                        "sentences": [content]  # Treating CSV data as one sentence
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"CSV parsing failed for {filepath}: {e}")

    def parse_xml(self, filepath, language=None):
        """
        Parse XML files.

        :param filepath: Path to the XML file.
        :param language: Unused here, kept for consistency.
        :return: Extracted content structure.
        """
        try:
            tree = ET.parse(filepath)
            root = tree.getroot()
            content = ET.tostring(root, encoding='unicode')
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "xml",
                        "content": content,
                        "sentences": [content]  # Treating XML data as one sentence
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"XML parsing failed for {filepath}: {e}")

    def parse_html(self, filepath, language=None):
        """
        Parse HTML files.

        :param filepath: Path to the HTML file.
        :param language: Unused here, kept for consistency.
        :return: Extracted content structure.
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as file:
                content = html.unescape(file.read())
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "html",
                        "content": content,
                        "sentences": content.split('. ')
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"HTML parsing failed for {filepath}: {e}")

    def parse_odt(self, filepath, language=None):
        """
        Parse OpenDocument Text (.odt) files.

        :param filepath: Path to the ODT file.
        :param language: Unused here, kept for consistency.
        :return: Extracted content structure.
        """
        try:
            textdoc = load(filepath)
            content = teletype.extractText(textdoc)
            return {
                "title": os.path.basename(filepath),
                "author": "Unknown",
                "chunks": [
                    {
                        "type": "document",
                                                "content": content,
                        "sentences": content.split('. ')
                    }
                ]
            }
        except Exception as e:
            raise ValueError(f"ODT parsing failed for {filepath}: {e}")

    def parse_directory(self, dirpath, language="eng"):
        """
        Parse all files in a directory recursively.

        :param dirpath: Path to the directory to parse.
        :param language: Language for OCR and speech recognition.
        :return: List of dictionaries containing filename and parsed content structure.
        """
        results = []
        try:
            for root, _, files in os.walk(dirpath):
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        content = self.parse(file_path, language=language)
                        results.append({"filename": file, "content": content})
                    except Exception as e:
                        logger.warning(f"Failed to parse {file_path}: {e}")
        except Exception as e:
            logger.error(f"Error parsing directory {dirpath}: {e}")
            raise
        return results

    def parse_generic(self, filepath, language=None):
        """
        Generic method to handle file types not explicitly supported.

        :param filepath: Path to the file.
        :param language: Unused here, kept for consistency.
        :return: A dictionary indicating the file type is not supported for content extraction.
        """
        _, file_extension = os.path.splitext(filepath)
        return {
            "title": os.path.basename(filepath),
            "author": "Unknown",
            "chunks": [
                {
                    "type": "unsupported",
                    "content": f"File type {file_extension} is not supported for content extraction.",
                    "sentences": [f"File type {file_extension} is not supported for content extraction."]
                }
            ]
        }

    def parse_and_store(self, filepath, mime_type=None, language="eng"):
        """
        Parse the file, structure the content, and store it using MemoryEngine.

        :param filepath: Path to the file.
        :param mime_type: MIME type of the file.
        :param language: Language for OCR or transcription.
        :return: Parsed content structure.
        """
        try:
            # Parse the file to get structured content
            parsed_content = self.parse(filepath, mime_type, language)
            
            # Use the new MemoryEngine method to process and store the content
            content_id = self.memory_engine.upload_and_process_content(
                file_path=filepath,
                content_type=mime_type if mime_type else mimetypes.guess_type(filepath)[0],
                title=parsed_content["title"],
                author=parsed_content["author"],
                metadata={
                    "type": "parsed_content", 
                    "source_file": filepath,
                    "mime_type": mime_type if mime_type else mimetypes.guess_type(filepath)[0]
                }
            )
            
            logger.info(f"Content from {filepath} was parsed and stored with ID: {content_id}")
            return parsed_content
        except Exception as e:
            logger.error(f"Failed to parse and store file {filepath}. Error details: {str(e)}")
            raise